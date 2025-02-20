import cv2
import time
import re
import os
import logging
import numpy as np
from datetime import datetime, timedelta
import pytesseract
from roboflow import Roboflow
import win32com.client as win32
from openpyxl.drawing.image import Image
import openpyxl
from PIL import Image as PILImage
# Configuração do logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

class ExcelControleProd:
    def __init__(self, arquivo_excel, sheet_name='Sheet1'):
        self.arquivo_excel = arquivo_excel
        self.sheet_name = sheet_name
        self.colunas = [
            'Sala produtiv',
            'Local',
            'Horário',
            'im com pred',
            'im sem predição'
        ]
        self.setup_logging()
    
    def setup_logging(self):
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('controle_producao.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def redimensionar_imagem_excel(self, img_path):
        """Redimensiona a imagem para um tamanho adequado para o Excel"""
        try:
            # Carregar imagem
            img = PILImage.open(img_path)
            
            # Definir dimensões máximas desejadas para a célula do Excel
            max_width = 200  # pixels
            max_height = 150  # pixels
            
            # Calcular proporção
            ratio = min(max_width/float(img.size[0]), max_height/float(img.size[1]))
            
            # Calcular novas dimensões mantendo proporção
            new_width = int(img.size[0] * ratio)
            new_height = int(img.size[1] * ratio)
            
            # Redimensionar
            img_resized = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
            
            # Criar nome temporário para a imagem redimensionada
            temp_path = f"temp_{os.path.basename(img_path)}"
            img_resized.save(temp_path)
            
            return temp_path
            
        except Exception as e:
            self.logger.error(f"Erro ao redimensionar imagem para Excel: {str(e)}")
            return img_path

    def inserir_registro(self, dados, imagem_pred_path, imagem_sem_pred_path):
        temp_files = []  # Lista para controlar arquivos temporários
        try:
            # Criar novo workbook se não existir
            if not os.path.exists(self.arquivo_excel):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = self.sheet_name
                # Escrever cabeçalhos
                for col, header in enumerate(self.colunas, 1):
                    ws.cell(row=1, column=col, value=header)
            else:
                wb = openpyxl.load_workbook(self.arquivo_excel)
                ws = wb[self.sheet_name]

            # Encontrar próxima linha vazia
            next_row = ws.max_row + 1

            # Inserir dados do registro
            ws.cell(row=next_row, column=1, value=dados.get('Sala produtiv', ''))
            ws.cell(row=next_row, column=2, value=dados.get('Local', ''))
            ws.cell(row=next_row, column=3, value=dados.get('Horário', ''))

            # Ajustar altura da linha para as imagens redimensionadas
            ws.row_dimensions[next_row].height = 80  # Reduzido de 120 para 80

            # Processar imagens
            for col, img_path in [(4, imagem_pred_path), (5, imagem_sem_pred_path)]:
                if img_path and os.path.exists(img_path):
                    try:
                        # Redimensionar imagem
                        temp_path = self.redimensionar_imagem_excel(img_path)
                        temp_files.append(temp_path)
                        
                        # Adicionar imagem redimensionada
                        img = Image(temp_path)
                        img.anchor = f'{chr(64 + col)}{next_row}'
                        ws.add_image(img)
                    except Exception as e:
                        self.logger.error(f"Erro ao inserir imagem: {str(e)}")

            # Ajustar larguras das colunas
            ws.column_dimensions['A'].width = 15  # Sala produtiv
            ws.column_dimensions['B'].width = 15  # Local
            ws.column_dimensions['C'].width = 25  # Horário
            ws.column_dimensions['D'].width = 25  # im com pred (reduzido de 40 para 25)
            ws.column_dimensions['E'].width = 25  # im sem predição (reduzido de 40 para 25)

            # Salvar workbook
            wb.save(self.arquivo_excel)
            wb.close()

            self.logger.info("Registro inserido com sucesso")
            return True

        except Exception as e:
            self.logger.error(f"Erro ao inserir registro: {str(e)}")
            return False
            
        finally:
            # Limpar arquivos temporários
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except Exception as e:
                    self.logger.error(f"Erro ao remover arquivo temporário {temp_file}: {str(e)}")


class VideoProcessor:
    def __init__(self):
        # Configuração do Tesseract
        self.caminho_tesseract = r'Tesseract/tesseract.exe'
        pytesseract.pytesseract.tesseract_cmd = self.caminho_tesseract
        self.config = '-c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 --psm 6'
        
        # Inicialização do modelo Roboflow
        self.rf = Roboflow(api_key="zMSFgl1hKRNzI85AvY7i")
        self.project = self.rf.workspace("gruponc").project("summer-trainee-eswym")
        self.modelo = self.project.version(5).model
        
        # Inicialização de variáveis de instância
        self.riscos_sala = {}
        self.horario = None
        self.sala = None
        self.controle_movimento = False 
        self.salas_ativas = set()
        self.ultimo_movimento = {}
        self.lista_riscos = ['Sem mascara', 'Celular', 'Sem luva', 'Bota']
        self.dict_salas = {1: 'REV 150-01', 2: 'REV 500-01', 3: 'REV 500-02'}
        self.frame_counter = 0
        self.contagem_frames_sem_movimento = {}
        self.masks = {}
        self.resize_width = 800
        self.resize_height = 600
        self.caminho_tesseract = r'Tesseract/tesseract.exe'
        pytesseract.pytesseract.tesseract_cmd = self.caminho_tesseract
        self.config = '-c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 --psm 6'
        
        self.rf = Roboflow(api_key="zMSFgl1hKRNzI85AvY7i")
        self.project = self.rf.workspace("gruponc").project("summer-trainee-eswym")
        self.modelo = self.project.version(5).model
        

        # New variables for risk tracking
        self.ultima_deteccao_risco = {}  # {channel_id: {risco: timestamp}}
        self.tempo_sem_movimento = {}    # {channel_id: testam}
        self.riscos_pendentes_email = {} # {channel_id: [riscos]}
        # Excel controller
        self.excel_controller = ExcelControleProd('ControleProd.xlsx')

    def inicializar_camera(self, source_path):
        cap = cv2.VideoCapture(source_path)
        if not cap.isOpened():
            logging.error(f"Erro ao abrir câmera: {source_path}")
            return None
        return cap

    def detectar_movimento(self, frame, background_subtractor):
        try:
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            blur = cv2.GaussianBlur(gray, (21, 21), 0)
            mask = background_subtractor.apply(blur)
            _, thresh = cv2.threshold(mask, 200, 255, cv2.THRESH_BINARY)
            contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            return max(contours, key=cv2.contourArea) if contours else None
        except Exception as e:
            logging.error(f"Erro na detecção de movimento: {e}")
            return None

    def initialize_mask(self, frame_shape, channel_id):
        height, width = frame_shape[:2]
        mask = np.ones((height, width), dtype=np.uint8) * 255
        
        if channel_id == 1:
            cv2.rectangle(mask, (int(width//3), 0), (int(width//2), int(height//5)), 0, -1)
        elif channel_id == 2:
            cv2.rectangle(mask, (int(width//1.5), 0), (int(width//2), int(height//5)), 0, -1)
        elif channel_id == 3:
            cv2.rectangle(mask, (int(width-200), 50), (int(width), int(height-200)), 0, -1) # SALA REV 150-01
        return mask

    def leitura_data_hora(self, frame):
        try:
            x_start, x_end = int(self.resize_width/1.5), self.resize_width
            y_start, y_end = 15, 70
            roi = frame[y_start:y_end, x_start:x_end]
            if roi.size == 0:
                return None
            
            imagem_cinza = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
            _, imagem_binaria = cv2.threshold(imagem_cinza, 235, 255, cv2.THRESH_BINARY)
            texto_detectado = pytesseract.image_to_string(imagem_binaria, lang='eng', config=self.config)
            numeros = ''.join(re.findall(r'\d', texto_detectado))
            
            if len(numeros) == 14:
                return datetime.strptime(numeros, "%d%m%Y%H%M%S")
            return None
        except Exception as e:
            logging.error(f"Erro na leitura de data/hora: {e}")
            return None

    def leitura_sala(self, frame):
        try:
            x_start, x_end = 20, 220
            y_start, y_end = int(self.resize_height/1.14), self.resize_height-10
            roi = frame[y_start:y_end, x_start:x_end]
            if roi.size == 0:
                return None
            
            imagem_cinza = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
            _, imagem_binaria = cv2.threshold(imagem_cinza, 235, 255, cv2.THRESH_BINARY)
            texto_detectado = pytesseract.image_to_string(imagem_binaria, lang='eng', config=self.config)
            
            if texto_detectado[0:3] in ['REV', 'REY']:
                sala = texto_detectado
                if texto_detectado[0:3] == 'REY':
                    sala = 'REV ' + texto_detectado[4:]
                return sala.strip()
            return None
        except Exception as e:
            logging.error(f"Erro na leitura da sala: {e}")
            return None

    def salvar_imagem_risco(self, frame_anotado, frame_original, channel_id, classe, x, y, width, height, folga=0.2):
        try:
            x_start = max(x - int(width * folga), 0)
            y_start = max(y - int(height * folga), 0)
            x_end = min(x + width + int(width * folga), frame_anotado.shape[1])
            y_end = min(y + height + int(height * folga), frame_anotado.shape[0])

            pessoa_crop1 = frame_anotado[y_start:y_end, x_start:x_end]
            pessoa_crop2 = frame_original[y_start:y_end, x_start:x_end]
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nome_imagem = f'{channel_id}_{classe}_{timestamp}.png'
            
            os.makedirs('Infra/Anotado/', exist_ok=True)
            os.makedirs('Infra/Original/', exist_ok=True)
            
            caminho_imagem1 = os.path.join('Infra/Anotado/', nome_imagem)
            caminho_imagem2 = os.path.join('Infra/Original/', nome_imagem)
            
            cv2.imwrite(caminho_imagem1, pessoa_crop1)
            cv2.imwrite(caminho_imagem2, pessoa_crop2)
            
            return caminho_imagem1, caminho_imagem2
        except Exception as e:
            logging.error(f"Erro ao salvar imagem: {e}")
            return None, None
    def salvar_registro_excel(self, channel_id, caminhos_fotos, lista_deteccoes):
        try:
            if not all(caminhos_fotos):
                return
            
            caminho_original, caminho_anotado = caminhos_fotos
            
            dados = {
                'Sala produtiv': self.dict_salas.get(channel_id, f'Sala {channel_id}'),
                'Local': f'Camera {channel_id}',
                'Horário': self.horario.strftime('%Y-%m-%d %H:%M:%S') if self.horario else datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            
            self.excel_controller.inserir_registro(
                dados=dados,
                imagem_pred_path=caminho_anotado,
                imagem_sem_pred_path=caminho_original
            )
            
        except Exception as e:
            logging.error(f"Erro ao salvar registro no Excel: {e}")
    def envio_Email(self, caminhos_fotos, lista_deteccoes, horario, sala_operacao):
        try:
            if not all(caminhos_fotos):
                return
            
            # Enviar email
            caminho_original, caminho_anotado = caminhos_fotos
            outlook = win32.Dispatch('outlook.application')
            email = outlook.CreateItem(0)
            
            email.To = 't0047831@ems.com.br'
            email.Subject = 'Relatório de Infração Detectada'
            email.Body = f"""
            Prezados (as) gestores,

            Segue em anexo o relatório de infração detectada no período de {horario.strftime('%d/%m/%Y %H:%M:%S')}, 
            com detalhes sobre os desvios identificados nas salas produtivas.

            Resumo das Infrações:
            • Classificação da não conformidade: {', '.join(lista_deteccoes)}
            • Sala Produtiva Impactada: {sala_operacao}

            Atenciosamente,
            Monitoramento NOVAMED
            """

            for caminho in [caminho_original, caminho_anotado]:
                caminho_absoluto = os.path.abspath(caminho)
                if os.path.exists(caminho_absoluto):
                    email.Attachments.Add(caminho_absoluto)

            email.Send()
            
            # Salvar no Excel após enviar email
            channel_id = int(sala_operacao.split('-')[1]) if '-' in sala_operacao else None
            if channel_id:
                self.salvar_registro_excel(channel_id, caminhos_fotos, lista_deteccoes)
            
            logging.info('E-mail enviado e registro salvo com sucesso!')
            
        except Exception as e:
            logging.error(f'Erro ao enviar e-mail: {e}')

    def associar_riscos_pessoas(self, riscos, pessoas, frame, frame_original, channel_id):
        try:
            # Inicializa lista de riscos detectados neste frame
            riscos_detectados = []
            
            for risco in riscos:
                x_r, y_r, w_r, h_r, classe_r = risco
                
                # Verifica se este risco já foi detectado anteriormente
                if (channel_id in self.ultima_deteccao_risco and 
                    classe_r in self.ultima_deteccao_risco[channel_id]):
                    continue
                
                centro_risco = (x_r + w_r // 2, y_r + h_r // 2)
                pessoa_mais_proxima = None
                menor_distancia = float('inf')
                
                for pessoa in pessoas:
                    x_p, y_p, w_p, h_p = pessoa
                    centro_pessoa = (x_p + w_p // 2, y_p + h_p // 2)
                    distancia = np.sqrt((centro_risco[0] - centro_pessoa[0])**2 + 
                                      (centro_risco[1] - centro_pessoa[1])**2)
                    
                    if distancia < menor_distancia:
                        menor_distancia = distancia
                        pessoa_mais_proxima = pessoa

                if pessoa_mais_proxima:
                    x_p, y_p, w_p, h_p = pessoa_mais_proxima
                    riscos_detectados.append((classe_r, (x_p, y_p, w_p, h_p)))
            
            # Se há novos riscos detectados
            if riscos_detectados:
                # Inicializa dicionários se necessário
                if channel_id not in self.ultima_deteccao_risco:
                    self.ultima_deteccao_risco[channel_id] = {}
                if channel_id not in self.riscos_pendentes_email:
                    self.riscos_pendentes_email[channel_id] = []
                
                # Processa todos os riscos detectados
                for classe_r, coords in riscos_detectados:
                    x_p, y_p, w_p, h_p = coords
                    if classe_r not in self.ultima_deteccao_risco[channel_id]:
                        self.ultima_deteccao_risco[channel_id][classe_r] = datetime.now()
                        self.riscos_pendentes_email[channel_id].append(classe_r)
                
                # Salva imagem e envia email apenas uma vez para todos os riscos
                if self.riscos_pendentes_email[channel_id]:
                    caminhos = self.salvar_imagem_risco(frame, frame_original, channel_id,
                                                      "_".join(self.riscos_pendentes_email[channel_id]),
                                                      x_p, y_p, w_p, h_p)
                    
                    if all(caminhos):
                        self.envio_Email(caminhos, self.riscos_pendentes_email[channel_id],
                                       self.horario, self.dict_salas.get(channel_id))
                        self.riscos_pendentes_email[channel_id] = []  # Limpa a lista após enviar
                
        except Exception as e:
            logging.error(f"Erro na associação de riscos: {e}")

    def realizar_predicao(self, frame, frame_original, channel_id):
        try:
            dados = self.modelo.predict(frame, confidence=0.5, overlap=0.5).json()#self.modelo.predict(frame, confidence=0.6, overlap=0.3).json()
            pessoas = []
            riscos = []

            for pred in dados['predictions']:
                x = int(pred['x'] - pred['width'] / 2)
                y = int(pred['y'] - pred['height'] / 2)
                width = int(pred['width'])
                height = int(pred['height'])
                classe = pred['class'].strip()
                
                # Verificar limites do frame
                x = max(0, min(x, frame.shape[1]))
                y = max(0, min(y, frame.shape[0]))
                width = min(width, frame.shape[1] - x)
                height = min(height, frame.shape[0] - y)
                
                if width <= 0 or height <= 0:
                    continue
                
                cor = (0, 255, 0)  # Verde para pessoa
                if classe == 'Pessoa':
                    pessoas.append((x, y, width, height))
                elif classe in self.lista_riscos:
                    cor = (0, 0, 255)  # Vermelho para riscos
                    riscos.append((x, y, width, height, classe))
                
                cv2.rectangle(frame, (x, y), (x + width, y + height), cor, 1)
                cv2.putText(frame, classe, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, cor, 1)

            if not pessoas and channel_id in self.riscos_sala:
                self.riscos_sala[channel_id] = []
            elif riscos:
                self.associar_riscos_pessoas(riscos, pessoas, frame, frame_original, channel_id)

        except Exception as e:
            logging.error(f"Erro na realização da predição: {e}")
    def verificar_tempo_sem_movimento(self, channel_id):
        """Verifica se passou 10  minutos sem movimento na sala"""
        if channel_id in self.tempo_sem_movimento:
            tempo_passado = datetime.now() - self.tempo_sem_movimento[channel_id]
            if tempo_passado > timedelta(minutes=10):
                # Reseta os riscos detectados após 2 minutos sem movimento
                if channel_id in self.ultima_deteccao_risco:
                    self.ultima_deteccao_risco[channel_id] = {}
                return True
        return False
    
    def processar_frame(self, channel_id, frame, background_subtractor):
        try:
            if frame is None or frame.size == 0:
                return np.zeros((self.resize_height, self.resize_width, 3), dtype=np.uint8)

            frame = cv2.resize(frame, (self.resize_width, self.resize_height))
            frame_original = frame.copy()
            
            if channel_id not in self.masks:
                self.masks[channel_id] = self.initialize_mask(frame.shape, channel_id)
            
            masked_frame = cv2.bitwise_and(frame, frame, mask=self.masks[channel_id])
            
            if self.frame_counter % 50 == 0:
                hora = self.leitura_data_hora(frame)
                if hora is not None:
                    self.horario = hora
                
                motion_detected = self.detectar_movimento(masked_frame, background_subtractor)
                if motion_detected is not None:
                    area = cv2.contourArea(motion_detected)
                    if area > 10000:
                        self.ultimo_movimento[channel_id] = datetime.now()
                        self.tempo_sem_movimento[channel_id] = datetime.now()
                        self.contagem_frames_sem_movimento[channel_id] = 0
                        self.controle_movimento = True
                        
                        if channel_id not in self.salas_ativas:
                            self.salas_ativas.add(channel_id)
                    else:
                        self.contagem_frames_sem_movimento[channel_id] = \
                            self.contagem_frames_sem_movimento.get(channel_id, 0) + 1
                        
                        # Atualiza o tempo sem movimento
                        if channel_id not in self.tempo_sem_movimento:
                            self.tempo_sem_movimento[channel_id] = datetime.now()
                else:
                    # Se não há movimento, verifica o tempo
                    self.verificar_tempo_sem_movimento(channel_id)
            
            if self.controle_movimento:
                if self.frame_counter % 50 == 0 and channel_id in self.salas_ativas:
                    self.realizar_predicao(masked_frame, frame_original, channel_id)
                    
                    if channel_id not in self.dict_salas:
                        sala = self.leitura_sala(frame)
                        if sala:
                            self.dict_salas[channel_id] = sala
                            self.sala = sala

            return masked_frame

        except Exception as e:
            logging.error(f"Erro no processamento do frame: {e}")
            return np.zeros((self.resize_height, self.resize_width, 3), dtype=np.uint8)
        
    def processar_videos(self, video_sources):
        try:
            background_subtractors = {
                channel_id: cv2.createBackgroundSubtractorMOG2(history=50, varThreshold=50, detectShadows=False)
                for channel_id in video_sources
            }
            
            caps = {channel_id: self.inicializar_camera(source) 
                   for channel_id, source in video_sources.items()}

            # Configuração do layout do mosaico
            n_cameras = len(video_sources)
            rows = 2 if n_cameras > 2 else 1
            cols = (n_cameras + 1) // 2
            
            window_name = "Camera Mosaic"
            cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
            
            mosaic_width = cols * self.resize_width
            mosaic_height = rows * self.resize_height
            cv2.resizeWindow(window_name, mosaic_width, mosaic_height)

            while True:
                mosaic = np.zeros((mosaic_height, mosaic_width, 3), dtype=np.uint8)
                
                for idx, (channel, cap) in enumerate(caps.items()):
                    if cap is None:
                        continue
                    
                    ret, frame = cap.read()
                    if not ret:
                        logging.error(f"Falha ao ler frame da câmera {channel}")
                        caps[channel] = self.inicializar_camera(video_sources[channel])
                        continue

                    frame = cv2.resize(frame, (self.resize_width, self.resize_height))
                    processed_frame = self.processar_frame(channel, frame, background_subtractors[channel])
                    
                    row = idx // cols
                    col = idx % cols
                    y_start = row * self.resize_height
                    y_end = (row + 1) * self.resize_height
                    x_start = col * self.resize_width
                    x_end = (col + 1) * self.resize_width
                    
                    mosaic[y_start:y_end, x_start:x_end] = processed_frame

                cv2.imshow(window_name, mosaic)
                self.frame_counter += 1

                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break

        except Exception as e:
            logging.error(f"Erro no processamento de vídeos: {e}")
        finally:
            for cap in caps.values():
                if cap:
                    cap.release()
            cv2.destroyAllWindows()

if __name__ == "__main__":
    try:
        video_sources = {
            1: "SALA_REV.mp4",
            2: "SALA_REV_500.mp4", 
            3: 'SALA_REV150.mp4'        }
        processor = VideoProcessor()
        processor.processar_videos(video_sources)
    except Exception as e:
        logging.error(f"Erro na execução principal: {e}")